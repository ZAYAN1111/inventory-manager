import os
from datetime import datetime
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QFileDialog,
    QMessageBox, QComboBox, QLabel, QDateEdit,
)
from PySide6.QtCore import QDate
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class SalesLog(QDialog):
    """Read-only table of past sales with date filtering and export."""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Sales Log")
        self.resize(1100, 600)

        layout = QVBoxLayout(self)

        # Filter row 1: Year/Month dropdowns
        filter_row1 = QHBoxLayout()
        filter_row1.addWidget(QLabel("Filter by Date:"))
        
        self.year_combo = QComboBox()
        self.year_combo.addItem("All Years")
        self.month_combo = QComboBox()
        self.month_combo.addItem("All Months")
        
        # Populate year/month from sales data
        sales = db.sales()
        years = set()
        months = set()
        for sale in sales:
            date_str = sale[4]  # sale_time
            try:
                date_obj = datetime.fromisoformat(date_str)
                years.add(date_obj.year)
                months.add(date_obj.month)
            except:
                pass
        
        self.year_combo.addItems(sorted([str(y) for y in years]))
        month_names = ["January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems([month_names[m-1] for m in sorted(months)])
        
        filter_row1.addWidget(self.year_combo)
        filter_row1.addWidget(self.month_combo)
        
        apply_date_btn = QPushButton("🔍 Apply Quick Filter")
        apply_date_btn.clicked.connect(self._apply_quick_date_filter)
        filter_row1.addWidget(apply_date_btn)
        
        filter_row1.addStretch()
        layout.addLayout(filter_row1)

        # Filter row 2: Date range picker
        filter_row2 = QHBoxLayout()
        filter_row2.addWidget(QLabel("Or pick date range:"))
        
        filter_row2.addWidget(QLabel("From:"))
        self.from_date = QDateEdit()
        self.from_date.setDate(QDate(2020, 1, 1))
        filter_row2.addWidget(self.from_date)
        
        filter_row2.addWidget(QLabel("To:"))
        self.to_date = QDateEdit()
        self.to_date.setDate(QDate.currentDate())
        filter_row2.addWidget(self.to_date)
        
        apply_range_btn = QPushButton("📅 Apply Range Filter")
        apply_range_btn.clicked.connect(self._apply_range_filter)
        filter_row2.addWidget(apply_range_btn)
        
        clear_btn = QPushButton("♻️ Clear Filters")
        clear_btn.clicked.connect(self._clear_filters)
        filter_row2.addWidget(clear_btn)
        
        filter_row2.addStretch()
        layout.addLayout(filter_row2)

        # Table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        rows = db.sales()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Item", "Brand", "Model", "Qty", "Date", "Customer"])
        self.refresh_table(rows)
        layout.addWidget(self.table)

        # Export buttons
        button_row = QHBoxLayout()
        button_row.addStretch()

        pdf_btn = QPushButton("📄 Export as PDF")
        pdf_btn.clicked.connect(self._export_pdf)
        button_row.addWidget(pdf_btn)

        excel_btn = QPushButton("📊 Export as Excel")
        excel_btn.clicked.connect(self._export_excel)
        button_row.addWidget(excel_btn)

        layout.addLayout(button_row)
        
        self.current_rows = rows

    def refresh_table(self, rows):
        """Populate table with sales."""
        self.current_rows = rows
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                self.table.setItem(r, c, QTableWidgetItem(str(value)))

    def _apply_quick_date_filter(self):
        """Filter by year and/or month."""
        year_text = self.year_combo.currentText()
        month_text = self.month_combo.currentText()
        
        sales = self.db.sales()
        filtered = []
        
        month_names = ["January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December"]
        month_map = {m: i+1 for i, m in enumerate(month_names)}
        
        for sale in sales:
            date_str = sale[4]
            try:
                date_obj = datetime.fromisoformat(date_str)
                
                year_match = (year_text == "All Years") or (str(date_obj.year) == year_text)
                month_match = (month_text == "All Months") or (date_obj.month == month_map.get(month_text, 0))
                
                if year_match and month_match:
                    filtered.append(sale)
            except:
                pass
        
        self.refresh_table(filtered)

    def _apply_range_filter(self):
        """Filter by date range."""
        from_date = self.from_date.date().toPython()
        to_date = self.to_date.date().toPython()
        
        sales = self.db.sales()
        filtered = []
        
        for sale in sales:
            date_str = sale[4]
            try:
                date_obj = datetime.fromisoformat(date_str).date()
                if from_date <= date_obj <= to_date:
                    filtered.append(sale)
            except:
                pass
        
        self.refresh_table(filtered)

    def _clear_filters(self):
        """Show all sales."""
        self.year_combo.setCurrentIndex(0)
        self.month_combo.setCurrentIndex(0)
        self.refresh_table(self.db.sales())

    def _export_pdf(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Sales Log as PDF", "", "PDF Files (*.pdf)"
        )
        if not path:
            return

        if not path.endswith(".pdf"):
            path += ".pdf"

        try:
            rows = self.db.sales()
            headers = ["Item", "Brand", "Model", "Qty", "Date", "Customer"]
            data = [headers] + [[str(v) for v in row] for row in rows]

            doc = SimpleDocTemplate(
                path, pagesize=landscape(letter), topMargin=20, bottomMargin=20
            )
            table = Table(data, colWidths=[90, 80, 80, 50, 120, 90])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
                    ]
                )
            )

            story = [
                Paragraph(f"<h2>Sales Log Export</h2>", getSampleStyleSheet()["Heading2"]),
                Spacer(1, 12),
                table,
                Spacer(1, 12),
                Paragraph(
                    f"<i>Exported on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>",
                    getSampleStyleSheet()["Normal"],
                ),
            ]
            doc.build(story)
            QMessageBox.information(self, "Success", f"Sales log exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export PDF:\n{str(e)}")

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Sales Log as Excel", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Log"

            headers = ["Item", "Brand", "Model", "Qty", "Date", "Customer"]
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_font = Font(bold=True, color="ffffff")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            rows = self.db.sales()
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 22
            ws.column_dimensions["F"].width = 18

            wb.save(path)
            QMessageBox.information(self, "Success", f"Sales log exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel:\n{str(e)}")
