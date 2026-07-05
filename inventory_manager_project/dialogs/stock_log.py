from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QPushButton,
    QFileDialog, QMessageBox, QLabel, QLineEdit, QComboBox,
)
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class StockLog(QDialog):
    """Stock inventory log table with quantity filtering and export.

    `items` should be the items currently shown on the results page
    (i.e. whatever the active search/filter produced) -- NOT the full
    inventory. This keeps the log/export scoped to what the person is
    actually looking at.
    """

    def __init__(self, items, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Stock Log")
        self.resize(1100, 600)

        layout = QVBoxLayout(self)

        # Filter row
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Stock Quantity:"))
        
        self.operator_combo = QComboBox()
        self.operator_combo.addItems([">=", "<=", "=", ">", "<"])
        self.operator_combo.setMaximumWidth(60)
        filter_layout.addWidget(self.operator_combo)
        
        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("Enter quantity")
        self.quantity_input.setMaximumWidth(100)
        filter_layout.addWidget(self.quantity_input)
        
        filter_btn = QPushButton("🔍 Filter")
        filter_btn.clicked.connect(self._apply_filter)
        filter_layout.addWidget(filter_btn)
        
        reset_filter_btn = QPushButton("♻️ Reset Filter")
        reset_filter_btn.clicked.connect(self._reset_filter)
        filter_layout.addWidget(reset_filter_btn)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Name", "Brand", "Model", "Origin", "Stock", "Minimum", "Photo"])
        self.all_items = items  # scoped to whatever was passed in, not the whole DB
        self._populate_table(self.all_items)
        layout.addWidget(self.table)

        # Export buttons
        export_layout = QHBoxLayout()
        export_layout.addStretch()

        pdf_btn = QPushButton("📄 Export as PDF")
        pdf_btn.clicked.connect(self._export_pdf)
        export_layout.addWidget(pdf_btn)

        excel_btn = QPushButton("📊 Export as Excel")
        excel_btn.clicked.connect(self._export_excel)
        export_layout.addWidget(excel_btn)

        layout.addLayout(export_layout)

    def _reset_to_all(self):
        self._populate_table(self.all_items)

    def _populate_table(self, items):
        self.table.setRowCount(len(items))
        for r, item in enumerate(items):
            item_id, name, brand, model, origin, stock, minimum, photo_path = item
            self.table.setItem(r, 0, QTableWidgetItem(str(name)))
            self.table.setItem(r, 1, QTableWidgetItem(str(brand or "")))
            self.table.setItem(r, 2, QTableWidgetItem(str(model or "")))
            self.table.setItem(r, 3, QTableWidgetItem(str(origin or "")))
            self.table.setItem(r, 4, QTableWidgetItem(str(stock)))
            self.table.setItem(r, 5, QTableWidgetItem(str(minimum)))
            self.table.setItem(r, 6, QTableWidgetItem("✓" if photo_path else ""))

    def _apply_filter(self):
        try:
            operator = self.operator_combo.currentText()
            qty = int(self.quantity_input.text())
            
            filtered = []
            for item in self.all_items:
                stock = item[5]
                if operator == ">=":
                    if stock >= qty:
                        filtered.append(item)
                elif operator == "<=":
                    if stock <= qty:
                        filtered.append(item)
                elif operator == "=":
                    if stock == qty:
                        filtered.append(item)
                elif operator == ">":
                    if stock > qty:
                        filtered.append(item)
                elif operator == "<":
                    if stock < qty:
                        filtered.append(item)
            
            self._populate_table(filtered)
        except ValueError:
            QMessageBox.warning(self, "Error", "Please enter a valid number")

    def _reset_filter(self):
        self.quantity_input.setText("")
        self._reset_to_all()

    def _export_pdf(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Stock Log as PDF", "", "PDF Files (*.pdf)"
        )
        if not path:
            return

        if not path.endswith(".pdf"):
            path += ".pdf"

        try:
            rows = self.all_items
            headers = ["Name", "Brand", "Model", "Origin", "Stock", "Minimum"]
            data = [headers] + [[str(v) for v in row[1:7]] for row in rows]

            doc = SimpleDocTemplate(
                path, pagesize=landscape(letter), topMargin=20, bottomMargin=20
            )
            table = Table(data, colWidths=[90, 80, 80, 80, 70, 70])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#27ae60")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
                    ]
                )
            )

            story = [
                Paragraph(f"<h2>Stock Inventory Log</h2>", getSampleStyleSheet()["Heading2"]),
                Spacer(1, 12),
                table,
                Spacer(1, 12),
                Paragraph(
                    f"<i>Exported on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>",
                    getSampleStyleSheet()["Normal"],
                ),
            ]
            doc.build(story)
            QMessageBox.information(self, "Success", f"Stock log exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export PDF:\n{str(e)}")

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Stock Log as Excel", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        if not path.endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Log"

            headers = ["Name", "Brand", "Model", "Origin", "Stock", "Minimum"]
            header_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            header_font = Font(bold=True, color="ffffff")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            rows = self.all_items
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row[1:7], 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 12

            wb.save(path)
            QMessageBox.information(self, "Success", f"Stock log exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel:\n{str(e)}")
